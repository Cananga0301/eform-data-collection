"""
T4 - Daily progress report and dashboard metrics.

The dashboard and downloadable reports share the same aggregation logic.
All totals exclude deactivated segments. Time-based progress uses first_seen_at.
"""

import logging
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill
from sqlalchemy import func, cast, Date as SADate
from sqlalchemy.orm import joinedload

from config import BRANCH_ALERT_DAYS, ETA_WINDOW_DAYS
from src.models.eform_models import Assignment, CollectedRecord, Segment, UnmappedRecord
from src.repository.eform_repository import EformRepository
from src.utils.text import normalize

logger = logging.getLogger(__name__)

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RED_FILL = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
STATUS_NOT_STARTED = normalize('Chưa bắt đầu')
STATUS_IN_PROGRESS = normalize('Đang thu thập')
STATUS_ENOUGH_POSITIONS = normalize('Đủ vị trí')
STATUS_ERROR = normalize('Dữ liệu sai hoặc lỗi')
STATUS_COMPLETED = normalize('Hoàn thành')

STATUS_ORDER = [
    STATUS_NOT_STARTED,
    STATUS_IN_PROGRESS,
    STATUS_ENOUGH_POSITIONS,
    STATUS_ERROR,
    STATUS_COMPLETED,
]
STATUS_LABELS = {
    STATUS_NOT_STARTED: 'Số đoạn đường chưa bắt đầu',
    STATUS_IN_PROGRESS: 'Số đoạn đường đang thu thập',
    STATUS_ENOUGH_POSITIONS: 'Số đoạn đường đủ vị trí',
    STATUS_ERROR: 'Số đoạn đường dữ liệu sai hoặc lỗi',
    STATUS_COMPLETED: 'Số đoạn đường hoàn thành',
}


class ReporterService:
    def __init__(self, repository: EformRepository):
        self.repo = repository

    def get_dashboard_metrics(
        self,
        tinh_thanh: str = None,
        xa_phuong: str = None,
    ) -> dict:
        return self.get_dashboard_data(
            tinh_thanh=tinh_thanh,
            xa_phuong=xa_phuong,
        )['metrics']

    def get_dashboard_data(
        self,
        tinh_thanh: str = None,
        xa_phuong: str = None,
    ) -> dict:
        now = datetime.now(timezone.utc)
        recent_cutoff = now - timedelta(days=BRANCH_ALERT_DAYS)
        eta_cutoff = now - timedelta(days=ETA_WINDOW_DAYS)

        with self.repo.session_scope() as session:
            segments = self._get_filtered_segments(session, tinh_thanh, xa_phuong)
            segment_rows = self._build_segment_rows(session, segments, recent_cutoff)
            seg_ids = [segment.id for segment in segments]

            total_needed = sum(row['needed'] for row in segment_rows)
            total_collected = sum(row['collected'] for row in segment_rows)
            pct_complete = round(total_collected / total_needed * 100, 1) if total_needed else 0.0
            eta = self._compute_eta(session, seg_ids, total_needed, total_collected, eta_cutoff)

            return {
                'metrics': {
                    'total_needed': total_needed,
                    'total_collected': total_collected,
                    'pct_complete': pct_complete,
                    'eta': eta,
                },
                'status_summary': self._build_status_summary(segment_rows),
                'status_counts': self._build_status_counts(segment_rows),
                'overview': self._build_breakdown_rows(segment_rows),
                'branch_activity': self._build_branch_activity_rows(segment_rows),
                'white_zones': self._build_white_zone_rows(segment_rows),
                'employee_stats': self._build_employee_stats_rows(session, segments, segment_rows),
                'recent_days': BRANCH_ALERT_DAYS,
            }

    def export_dashboard_excel(
        self,
        tinh_thanh: str = None,
        xa_phuong: str = None,
    ) -> bytes:
        dashboard = self.get_dashboard_data(tinh_thanh=tinh_thanh, xa_phuong=xa_phuong)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._build_dashboard_summary_sheet(wb, dashboard)
        self._build_dashboard_overview_sheet(wb, dashboard)
        self._build_dashboard_branch_activity_sheet(wb, dashboard)
        self._build_dashboard_white_zones_sheet(wb, dashboard)
        self._build_dashboard_employee_stats_sheet(wb, dashboard)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_daily_report(self, report_date: Optional[date] = None) -> bytes:
        """Generate 3-sheet Excel report. Returns raw bytes."""
        if report_date is None:
            report_date = date.today()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        with self.repo.session_scope() as session:
            recent_cutoff = datetime.combine(
                report_date - timedelta(days=BRANCH_ALERT_DAYS),
                datetime.min.time(),
            ).replace(tzinfo=timezone.utc)
            segments = self._get_filtered_segments(session)
            segment_rows = self._build_segment_rows(session, segments, recent_cutoff)
            breakdown_rows = self._build_breakdown_rows(segment_rows)

            self._build_sheet1(wb, breakdown_rows)
            self._build_sheet2(wb, segment_rows)
            self._build_sheet3(wb, session)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _get_filtered_segments(self, session, tinh_thanh: str = None, xa_phuong: str = None) -> list[Segment]:
        q = session.query(Segment).options(joinedload(Segment.branch)).filter(Segment.is_active == True)
        if tinh_thanh:
            q = q.filter(Segment.tinh_thanh_norm == normalize(tinh_thanh))
        if xa_phuong:
            q = q.filter(Segment.xa_phuong_norm == normalize(xa_phuong))
        return q.order_by(Segment.tinh_thanh, Segment.xa_phuong, Segment.ten_duong, Segment.doan).all()

    def _build_segment_rows(self, session, segments: list[Segment], recent_cutoff: datetime) -> list[dict]:
        if not segments:
            return []

        seg_ids = [segment.id for segment in segments]
        assignment_rows = session.query(Assignment).options(joinedload(Assignment.branch)).filter(
            Assignment.segment_id.in_(seg_ids)
        ).all()
        assignment_map = {assignment.segment_id: assignment for assignment in assignment_rows}

        collected_counts = dict(
            session.query(CollectedRecord.segment_id, func.count(CollectedRecord.id))
            .filter(
                CollectedRecord.segment_id.in_(seg_ids),
                CollectedRecord.is_active == True,
            )
            .group_by(CollectedRecord.segment_id)
            .all()
        )

        recent_counts = dict(
            session.query(CollectedRecord.segment_id, func.count(CollectedRecord.id))
            .filter(
                CollectedRecord.segment_id.in_(seg_ids),
                CollectedRecord.is_active == True,
                CollectedRecord.first_seen_at >= recent_cutoff,
            )
            .group_by(CollectedRecord.segment_id)
            .all()
        )

        rows = []
        for segment in segments:
            assignment = assignment_map.get(segment.id)
            branch_name = self._resolve_branch_name(segment, assignment)
            needed = self._segment_needed(segment)
            collected = int(collected_counts.get(segment.id, 0))
            recent_new = int(recent_counts.get(segment.id, 0))
            missing = max(needed - collected, 0)
            pct_complete = round(collected / needed * 100, 1) if needed else 0.0

            rows.append({
                'segment_id': segment.id,
                'province': segment.tinh_thanh or 'Unknown',
                'ward': segment.xa_phuong or '',
                'road': segment.ten_duong or '',
                'segment': segment.doan or '',
                'group': segment.nhom or '?',
                'status': segment.trang_thai or 'Unknown',
                'branch': branch_name,
                'person': assignment.phu_trach if assignment and assignment.phu_trach else '',
                'deadline': assignment.deadline.isoformat() if assignment and assignment.deadline else '',
                'needed': needed,
                'collected': collected,
                'missing': missing,
                'pct_complete': pct_complete,
                'new_last_2d': recent_new,
            })

        return rows

    def _resolve_branch_name(self, segment: Segment, assignment: Optional[Assignment]) -> str:
        if assignment and assignment.branch:
            return assignment.branch.name
        if segment.branch:
            return segment.branch.name
        return 'Unassigned'

    def _segment_needed(self, segment: Segment) -> int:
        return (
            (segment.so_can_vt1 or 0) +
            (segment.so_can_vt2 or 0) +
            (segment.so_can_vt3 or 0) +
            (segment.so_can_vt4 or 0)
        )

    def _build_status_summary(self, segment_rows: list[dict]) -> list[dict]:
        counts = Counter(normalize(row['status']) for row in segment_rows)

        def sort_key(status_name: str):
            try:
                return (0, STATUS_ORDER.index(status_name))
            except ValueError:
                return (1, status_name)

        return [
            {
                'status': STATUS_LABELS.get(status_name, status_name),
                'segments': counts[status_name],
            }
            for status_name in sorted(counts.keys(), key=sort_key)
        ]

    def _build_status_counts(self, segment_rows: list[dict]) -> dict:
        counts = Counter(normalize(row['status']) for row in segment_rows)
        return {
            'not_started': counts.get(STATUS_NOT_STARTED, 0),
            'in_progress': counts.get(STATUS_IN_PROGRESS, 0),
            'enough_positions': counts.get(STATUS_ENOUGH_POSITIONS, 0),
            'error': counts.get(STATUS_ERROR, 0),
            'completed': counts.get(STATUS_COMPLETED, 0),
        }

    def _build_breakdown_rows(self, segment_rows: list[dict]) -> list[dict]:
        grouped = defaultdict(lambda: {'needed': 0, 'collected': 0, 'missing': 0, 'new_last_2d': 0})
        branch_new_totals = defaultdict(int)

        for row in segment_rows:
            key = (row['province'], row['branch'], row['group'])
            grouped[key]['needed'] += row['needed']
            grouped[key]['collected'] += row['collected']
            grouped[key]['missing'] += row['missing']
            grouped[key]['new_last_2d'] += row['new_last_2d']
            branch_new_totals[(row['province'], row['branch'])] += row['new_last_2d']

        rows = []
        for province, branch, group in sorted(grouped.keys()):
            values = grouped[(province, branch, group)]
            rows.append({
                'Province': province,
                'Branch': branch,
                'Group': group,
                'Needed': values['needed'],
                'Collected': values['collected'],
                'Missing': values['missing'],
                '% Complete': round(values['collected'] / values['needed'] * 100, 1) if values['needed'] else 0.0,
                'New (last 2d)': values['new_last_2d'],
                'Branch Alert': 'No new records' if branch_new_totals[(province, branch)] == 0 else '',
            })
        return rows

    def _build_branch_activity_rows(self, segment_rows: list[dict]) -> list[dict]:
        grouped = defaultdict(lambda: {'needed': 0, 'collected': 0, 'missing': 0, 'new_last_2d': 0, 'segments': 0})

        for row in segment_rows:
            key = (row['province'], row['branch'])
            grouped[key]['needed'] += row['needed']
            grouped[key]['collected'] += row['collected']
            grouped[key]['missing'] += row['missing']
            grouped[key]['new_last_2d'] += row['new_last_2d']
            grouped[key]['segments'] += 1

        rows = []
        for province, branch in sorted(grouped.keys()):
            values = grouped[(province, branch)]
            rows.append({
                'Province': province,
                'Branch': branch,
                'Segments': values['segments'],
                'Needed': values['needed'],
                'Collected': values['collected'],
                'Missing': values['missing'],
                '% Complete': round(values['collected'] / values['needed'] * 100, 1) if values['needed'] else 0.0,
                'New (last 2d)': values['new_last_2d'],
                'Alert': 'No new records' if values['new_last_2d'] == 0 else '',
            })
        return rows

    def _build_white_zone_rows(self, segment_rows: list[dict]) -> list[dict]:
        rows = [
            {
                'Segment ID': row['segment_id'],
                'Province': row['province'],
                'Ward / Zone': row['ward'],
                'Road': row['road'],
                'Segment': row['segment'],
                'Group': row['group'],
                'Branch': row['branch'],
                'Person': row['person'],
                'Deadline': row['deadline'],
                'Status': row['status'],
                'Missing': row['missing'],
            }
            for row in segment_rows
            if row['group'] in {'A', 'B'} and row['missing'] > 0
        ]
        rows.sort(key=lambda row: (-row['Missing'], row['Province'], row['Ward / Zone'], row['Road']))
        return rows

    def _build_employee_stats_rows(
        self,
        session,
        segments: list[Segment],
        segment_rows: list[dict],
    ) -> list[dict]:
        from zoneinfo import ZoneInfo
        seg_ids = [s.id for s in segments]
        if not seg_ids:
            return []

        VN_TZ = ZoneInfo('Asia/Ho_Chi_Minh')
        today_vn = datetime.now(VN_TZ).date()

        vn_date = cast(func.timezone('Asia/Ho_Chi_Minh', CollectedRecord.first_seen_at), SADate)

        before_deadline_counts = dict(
            session.query(CollectedRecord.segment_id, func.count(CollectedRecord.id))
            .join(Assignment, Assignment.segment_id == CollectedRecord.segment_id)
            .filter(
                CollectedRecord.segment_id.in_(seg_ids),
                CollectedRecord.is_active == True,
                Assignment.deadline.isnot(None),
                vn_date <= Assignment.deadline,
            )
            .group_by(CollectedRecord.segment_id)
            .all()
        )

        after_deadline_counts = dict(
            session.query(CollectedRecord.segment_id, func.count(CollectedRecord.id))
            .join(Assignment, Assignment.segment_id == CollectedRecord.segment_id)
            .filter(
                CollectedRecord.segment_id.in_(seg_ids),
                CollectedRecord.is_active == True,
                Assignment.deadline.isnot(None),
                vn_date > Assignment.deadline,
            )
            .group_by(CollectedRecord.segment_id)
            .all()
        )

        inactive_counts = dict(
            session.query(CollectedRecord.segment_id, func.count(CollectedRecord.id))
            .filter(
                CollectedRecord.segment_id.in_(seg_ids),
                CollectedRecord.is_active == False,
            )
            .group_by(CollectedRecord.segment_id)
            .all()
        )

        UNASSIGNED = '(Chưa phân công)'
        emp = defaultdict(lambda: {
            'branch': '', 'employee_display': '',
            'segments': 0, 'needed': 0, 'collected': 0,
            'before_deadline': 0, 'after_deadline': 0, 'no_deadline': 0,
            'overdue_open_segments': 0, 'error_segments': 0,
            'inactive_records': 0, 'new_last_nd': 0,
        })

        for row in segment_rows:
            person_raw = (row['person'] or '').strip()
            person_norm = normalize(person_raw) if person_raw else ''
            key = (row['branch'], person_norm)

            data = emp[key]
            data['branch'] = row['branch']
            if not data['employee_display']:
                data['employee_display'] = person_raw if person_raw else UNASSIGNED
            data['segments']         += 1
            data['needed']           += row['needed']
            data['collected']        += row['collected']
            data['new_last_nd']      += row['new_last_2d']
            data['inactive_records'] += int(inactive_counts.get(row['segment_id'], 0))

            if normalize(row['status']) == STATUS_ERROR:
                data['error_segments'] += 1

            data['before_deadline'] += int(before_deadline_counts.get(row['segment_id'], 0))
            data['after_deadline']  += int(after_deadline_counts.get(row['segment_id'], 0))

            if not row['deadline']:
                data['no_deadline'] += row['collected']

            if row['deadline'] and row['deadline'] < today_vn.isoformat() and row['missing'] > 0:
                data['overdue_open_segments'] += 1

        recent_label = f'New (last {BRANCH_ALERT_DAYS}d)'
        result = []
        for (_, __), data in emp.items():
            missing = max(data['needed'] - data['collected'], 0)
            pct = round(data['collected'] / data['needed'] * 100, 1) if data['needed'] else 0.0
            is_overdue = data['overdue_open_segments'] > 0
            is_idle = data['new_last_nd'] == 0 and missing > 0 and not is_overdue
            result.append({
                'Branch':                data['branch'],
                'Employee':              data['employee_display'],
                'Segments':              data['segments'],
                'Needed':                data['needed'],
                'Collected':             data['collected'],
                'Missing':               missing,
                '% Complete':            pct,
                'Before Deadline':       data['before_deadline'],
                'After Deadline':        data['after_deadline'],
                'Overdue Open Segments': data['overdue_open_segments'],
                'Error Segments':        data['error_segments'],
                recent_label:            data['new_last_nd'],
                'No Deadline':           data['no_deadline'],
                'Inactive Records':      data['inactive_records'],
                '_overdue':    is_overdue,
                '_idle':       is_idle,
                '_unassigned': data['employee_display'] == UNASSIGNED,
            })

        result.sort(key=lambda r: (-r['Overdue Open Segments'], -r['Missing'], r['% Complete']))
        return result

    def _build_sheet1(self, wb, overview_rows: list[dict]):
        ws = wb.create_sheet('Overview')
        headers = ['Province', 'Branch', 'Group', 'Total Needed', 'Collected', '% Done', 'New (last 2d)']
        ws.append(headers)

        for row in overview_rows:
            ws.append([
                row['Province'],
                row['Branch'],
                row['Group'],
                row['Needed'],
                row['Collected'],
                row['% Complete'],
                row['New (last 2d)'],
            ])
            if row['Branch Alert']:
                for cell in ws[ws.max_row]:
                    cell.fill = YELLOW_FILL

    def _build_sheet2(self, wb, segment_rows: list[dict]):
        ws = wb.create_sheet('White Zones')
        headers = ['Province', 'Ward/Zone', 'Road', 'Segment', 'Group', 'Branch', 'Person', 'Missing']
        ws.append(headers)

        for row in self._build_white_zone_rows(segment_rows):
            ws.append([
                row['Province'],
                row['Ward / Zone'],
                row['Road'],
                row['Segment'],
                row['Group'],
                row['Branch'],
                row['Person'],
                row['Missing'],
            ])

    def _build_sheet3(self, wb, session):
        ws = wb.create_sheet('Unmapped Records')
        headers = ['ID', 'Source Record ID', 'Reason', 'Raw Data Preview']
        ws.append(headers)

        unmapped = session.query(UnmappedRecord).filter_by(resolved=False).all()
        for row in unmapped:
            preview = str(row.raw_data)[:200] if row.raw_data else ''
            ws.append([row.id, row.source_record_id, row.reason, preview])

    def _build_dashboard_summary_sheet(self, wb, dashboard: dict):
        ws = wb.create_sheet('Summary')
        metrics = dashboard['metrics']

        ws.append(['Metric', 'Value'])
        ws.append(['Total Needed', metrics['total_needed']])
        ws.append(['Collected', metrics['total_collected']])
        ws.append(['% Complete', metrics['pct_complete']])
        ws.append(['ETA', metrics['eta']])
        ws.append([])
        ws.append(['Status', 'Segments'])
        for row in dashboard['status_summary']:
            ws.append([row['status'], row['segments']])

    def _build_dashboard_overview_sheet(self, wb, dashboard: dict):
        ws = wb.create_sheet('Dashboard Overview')
        headers = ['Province', 'Branch', 'Group', 'Needed', 'Collected', 'Missing', '% Complete', 'New (last 2d)', 'Branch Alert']
        ws.append(headers)

        for row in dashboard['overview']:
            ws.append([
                row['Province'],
                row['Branch'],
                row['Group'],
                row['Needed'],
                row['Collected'],
                row['Missing'],
                row['% Complete'],
                row['New (last 2d)'],
                row['Branch Alert'],
            ])
            if row['Branch Alert']:
                for cell in ws[ws.max_row]:
                    cell.fill = YELLOW_FILL

    def _build_dashboard_branch_activity_sheet(self, wb, dashboard: dict):
        ws = wb.create_sheet('Branch Activity')
        headers = ['Province', 'Branch', 'Segments', 'Needed', 'Collected', 'Missing', '% Complete', 'New (last 2d)', 'Alert']
        ws.append(headers)

        for row in dashboard['branch_activity']:
            ws.append([
                row['Province'],
                row['Branch'],
                row['Segments'],
                row['Needed'],
                row['Collected'],
                row['Missing'],
                row['% Complete'],
                row['New (last 2d)'],
                row['Alert'],
            ])
            if row['Alert']:
                for cell in ws[ws.max_row]:
                    cell.fill = YELLOW_FILL

    def _build_dashboard_white_zones_sheet(self, wb, dashboard: dict):
        ws = wb.create_sheet('White Zones')
        headers = ['Segment ID', 'Province', 'Ward / Zone', 'Road', 'Segment', 'Group', 'Branch', 'Person', 'Deadline', 'Status', 'Missing']
        ws.append(headers)

        for row in dashboard['white_zones']:
            ws.append([
                row['Segment ID'],
                row['Province'],
                row['Ward / Zone'],
                row['Road'],
                row['Segment'],
                row['Group'],
                row['Branch'],
                row['Person'],
                row['Deadline'],
                row['Status'],
                row['Missing'],
            ])

    def _build_dashboard_employee_stats_sheet(self, wb, dashboard: dict):
        ws = wb.create_sheet('Employee Stats')
        recent_days = dashboard['recent_days']
        recent_label = f'New (last {recent_days}d)'
        headers = [
            'Branch', 'Employee', 'Segments', 'Needed', 'Collected', 'Missing',
            '% Complete', 'Before Deadline', 'After Deadline', 'Overdue Open Segments',
            'Error Segments', recent_label, 'No Deadline', 'Inactive Records',
        ]
        ws.append(headers)

        for row in dashboard['employee_stats']:
            ws.append([
                row['Branch'],
                row['Employee'],
                row['Segments'],
                row['Needed'],
                row['Collected'],
                row['Missing'],
                row['% Complete'],
                row['Before Deadline'],
                row['After Deadline'],
                row['Overdue Open Segments'],
                row['Error Segments'],
                row[recent_label],
                row['No Deadline'],
                row['Inactive Records'],
            ])
            if row.get('_overdue'):
                for cell in ws[ws.max_row]:
                    cell.fill = RED_FILL
            elif row.get('_idle'):
                for cell in ws[ws.max_row]:
                    cell.fill = YELLOW_FILL

    def _compute_eta(
        self,
        session,
        seg_ids: list[int],
        total_needed: int,
        total_collected: int,
        window_start: datetime,
    ) -> str:
        remaining = total_needed - total_collected
        if remaining <= 0:
            return date.today().isoformat()

        if not seg_ids:
            return 'Unknown'

        new_in_window = session.query(func.count(CollectedRecord.id)).filter(
            CollectedRecord.segment_id.in_(seg_ids),
            CollectedRecord.is_active == True,
            CollectedRecord.first_seen_at >= window_start,
        ).scalar() or 0

        velocity = new_in_window / ETA_WINDOW_DAYS
        if velocity == 0:
            return 'Unknown'

        days_left = remaining / velocity
        return (date.today() + timedelta(days=days_left)).isoformat()
